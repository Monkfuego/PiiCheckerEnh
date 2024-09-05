import os
from docx import Document
import PyPDF2
from pdfminer.high_level import extract_text as pdfminer_extract_text
from bs4 import BeautifulSoup
import os
import requests
from bs4 import BeautifulSoup
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import numpy as np
from PIL import Image as PILImage

output_folder = "Imagefolder"

def read_text_from_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.txt':
        return read_text_from_txt(file_path)
    elif ext == '.docx':
        return read_text_from_docx(file_path)
    elif ext == '.pdf':
        return read_text_from_pdf(file_path)
    elif ext == '.html' or ext == '.htm':
        return read_text_from_html(file_path)   
    else:
        raise ValueError(f"Unsupported file extension: {ext}")

def read_text_from_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        return file.read()

def read_text_from_docx(file_path):
    doc = Document(file_path)
    text = '\n'.join([para.text for para in doc.paragraphs])
    return text

def read_text_from_pdf(file_path):
    text = ""
    try:
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfFileReader(file)
            for page_num in range(reader.getNumPages()):
                text += reader.getPage(page_num).extractText()
        if not text.strip():  
            text = pdfminer_extract_text(file_path)
    except Exception as e:
        text = pdfminer_extract_text(file_path)
    return text

def read_text_from_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')
        return soup.get_text()
    
######################### Images ########

def extract_images_from_pdf(pdf_path, output_folder):
    """Extract images from a PDF file."""
    doc = fitz.open(pdf_path)
    for i in range(len(doc)):
        page = doc.load_page(i)
        image_list = page.get_images(full=True)
        for img_index, img in enumerate(image_list):
            xref = img[0]
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]
            image_filename = f"{output_folder}/pdf_page_{i+1}_img_{img_index+1}.png"
            with open(image_filename, "wb") as img_file:
                img_file.write(image_bytes)
    print("PDF images extracted.")

def extract_images_from_excel(excel_path, output_folder):
    """Extract images from an Excel file."""
    wb = load_workbook(excel_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for image in ws._images:
            img = image.image
            image_filename = f"{output_folder}/excel_{sheet}_image.png"
            img.save(image_filename)
    print("Excel images extracted.")

def extract_images_from_html(html_path, output_folder):
    """Extract images from an HTML file."""
    with open(html_path, "r", encoding="utf-8") as file:
        soup = BeautifulSoup(file, "html.parser")
    
    images = soup.find_all("img")
    for img in images:
        img_url = img.get("src")
        if img_url:
            img_name = os.path.basename(img_url)
            img_path = os.path.join(output_folder, img_name)
            img_data = requests.get(img_url).content
            with open(img_path, "wb") as handler:
                handler.write(img_data)
    print("HTML images extracted.")



def preprocess_image(image_path):
    """Preprocess image for model prediction."""
    img = PILImage.open(image_path).convert('RGB')
    img = img.resize((224, 224))  # Resize to model's input size
    img_array = np.array(img) / 255.0  # Normalize image
    img_array = np.expand_dims(img_array, axis=0)  # Add batch dimension
    return img_array

def image_extraction(file_path):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    file_extension = os.path.splitext(file_path)[1].lower()
    if file_extension == '.pdf':
        extract_images_from_pdf(file_path, output_folder)
    elif file_extension == '.xlsx':
        extract_images_from_excel(file_path, output_folder)
    elif file_extension == '.html':
        extract_images_from_html(file_path, output_folder)
    else:
        print(f"Unsupported file type: {file_extension}")
